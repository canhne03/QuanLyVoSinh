from flask import Flask, flash, redirect, render_template, request, url_for, session, send_file, jsonify
from datetime import datetime
import re
import os
from io import BytesIO
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# import từ models (phiên bản bạn đã yêu cầu)
from models import (
    db,
    HoiVien,
    TaiKhoan,
    HoSoVoSinh,
    KyThi,
    DangKyKyThi,
    init_db,
    get_trinh_do,
    get_mau_dai,
)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = "change-me-please"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///app.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Khởi tạo DB
init_db(app)

ALLOWED_EXT = {".xlsx", ".xls", ".csv"}

# ====== Tài khoản admin mặc định ======
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "123456"


# ====== Hàm tính sắp tự nhiên A1 < A2 < A10 ======
def natural_keys(text):
    """Tách số và chữ để sort đúng kiểu A1 < A2 < A10"""
    if text is None:
        return []
    return [int(c) if c.isdigit() else c for c in re.split(r"(\d+)", str(text))]


# ====== Trang chính (form đăng nhập) ======
@app.route("/")
def home():
    return render_template("index.html")


# ====== Xử lý đăng nhập Admin ======
@app.route("/login/admin", methods=["POST"])
def login_admin():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session["role"] = "admin"
        return jsonify({"success": True, "redirect": url_for("admin_students")})
    else:
        return jsonify({"success": False, "error": "Sai tài khoản hoặc mật khẩu!"})


# ====== Đăng xuất ======
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


# ====== Trang quản trị (dashboard) ======
@app.route("/quantri")
def quantri():
    if session.get("role") != "admin":
        # chuyển về trang login (ở đây về home, vì login_admin là POST-only)
        return redirect(url_for("home"))

    students = HoSoVoSinh.query.order_by(HoSoVoSinh.vocotruyenid.desc()).limit(200).all()
    return render_template("quantri.html", current_page="students", students=students)


# ================= QUẢN LÝ CÁC TRANG KHÁC =================
@app.route("/admin/gioithieu")
def admin_gioithieu():
    if session.get("role") != "admin":
        return redirect(url_for("home"))
    return render_template("admin/gioithieu.html")


@app.route("/admin/tintuc")
def admin_tintuc():
    if session.get("role") != "admin":
        return redirect(url_for("home"))
    return render_template("admin/tintuc.html")


@app.route("/admin/lienket")
def admin_lienket():
    if session.get("role") != "admin":
        return redirect(url_for("home"))
    return render_template("admin/lienket.html")


@app.route("/admin/clb")
def admin_clb():
    if session.get("role") != "admin":
        return redirect(url_for("home"))
    return render_template("admin/caulacbo.html")


@app.route("/admin/tailieu")
def admin_tailieu():
    if session.get("role") != "admin":
        return redirect(url_for("home"))
    return render_template("admin/tailieu.html")


@app.route("/admin/taikhoan")
def admin_taikhoan():
    if session.get("role") != "admin":
        return redirect(url_for("home"))
    return render_template("admin/taikhoan.html")


# ================= QUẢN LÝ VÕ SINH =================
@app.route("/admin/students")
def admin_students():
    if session.get("role") != "admin":
        return redirect(url_for("home"))

    students = HoSoVoSinh.query.all()
    students = sorted(students, key=lambda s: natural_keys(s.vocotruyenid))
    return render_template("admin/vosinh.html", students=students)


# ====== Thêm võ sinh ======
@app.route("/admin/students/add", methods=["POST"])
def admin_students_add():
    if session.get("role") != "admin":
        return redirect(url_for("home"))

    code = request.form.get("vocotruyenid", "").strip()
    hovaten = request.form.get("hovaten", "").strip()
    namsinh = request.form.get("namsinh")
    gioitinh = request.form.get("gioitinh")
    capbac = request.form.get("capbac")
    donvi = request.form.get("donvi")
    thanhtich = request.form.get("thanhtich")

    if not code or not hovaten:
        flash("Mã số và Họ tên là bắt buộc!")
        return redirect(url_for("admin_students"))

    # kiểm tra trùng
    s = HoSoVoSinh.query.filter_by(vocotruyenid=code).first()
    if s:
        flash(f"Võ sinh mã {code} đã tồn tại!")
        return redirect(url_for("admin_students"))

    # tính trình độ + maudai (lưu relative filename, ví dụ 'cap1.jpg' hoặc 'phongtang.jpg')
    trinhdo = get_trinh_do(capbac)
    maudai_rel = get_mau_dai(capbac, mode="rel") if capbac else None

    s = HoSoVoSinh(
        vocotruyenid=code,
        hovaten=hovaten,
        namsinh=int(namsinh) if namsinh else None,
        gioitinh=gioitinh,
        donvi=donvi,
        capbac=capbac,
        trinhdo=trinhdo,
        maudai=maudai_rel,
        thanhtich=thanhtich,
    )
    db.session.add(s)
    db.session.commit()
    return redirect(url_for("admin_students"))


# ====== Import võ sinh từ file (từ thiết bị) ======
@app.route("/admin/students/import", methods=["GET", "POST"])
def admin_students_import():
    if session.get("role") != "admin":
        return redirect(url_for("home"))

    if request.method == "POST":
        f = request.files.get("file")
        if not f:
            flash("Chưa chọn file!")
            return redirect(url_for("admin_students_import"))

        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXT:
            flash("Định dạng file không hợp lệ! (chỉ hỗ trợ .xlsx, .xls, .csv)")
            return redirect(url_for("admin_students_import"))

        filename = secure_filename(f.filename)
        tmp_path = os.path.join(app.config["UPLOAD_FOLDER"], f"tmp_{filename}")
        f.save(tmp_path)

        # đọc file bằng pandas (import tại chỗ để tránh startup nặng)
        import pandas as pd

        try:
            if ext == ".csv":
                df = pd.read_csv(tmp_path, dtype=str)
            else:
                # pandas sẽ tự chọn engine nếu có; openpyxl cho xlsx
                df = pd.read_excel(tmp_path, dtype=str)
        except Exception as e:
            flash(f"Lỗi khi đọc file: {e}")
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            return redirect(url_for("admin_students_import"))

        try:
            os.remove(tmp_path)
        except Exception:
            pass

        if df is None or df.empty:
            flash("File không có dữ liệu!")
            return redirect(url_for("admin_students_import"))

        # Chuẩn hóa tên cột: lower + strip
        df.columns = [str(c).strip().lower() for c in df.columns]

        def get_val(row, candidates):
            for k in candidates:
                # k có thể là chuỗi (đã lowercase), thử cả dạng có dấu và không dấu
                if k in row and pd.notna(row[k]):
                    return row[k]
            return None

        import pandas as pd  # đảm bảo pd tồn tại trong scope

        for _, raw_row in df.iterrows():
            # raw_row is a Series with lowercase keys
            # Thử các tên cột khả dĩ
            code = (
                get_val(raw_row, ["vocotruyenid", "mã số", "mã", "ma", "ms", "id", "mã_số", "vocotruyen"])
                or None
            )
            if not code:
                # fallback: cột đầu tiên
                first_col = df.columns[0]
                code = raw_row.get(first_col)

            if not code or str(code).strip() == "":
                continue
            code = str(code).strip()

            hovaten = get_val(raw_row, ["hovaten", "họ tên", "hoten", "ten", "họ_tên"]) or ""
            namsinh = get_val(raw_row, ["namsinh", "năm sinh", "ns"]) or None
            gioitinh = get_val(raw_row, ["gioitinh", "giới tính", "gioi tinh", "gt"]) or None
            capbac = get_val(raw_row, ["capbac", "cấp bậc", "cap", "capbac"]) or None
            donvi = get_val(raw_row, ["donvi", "đơn vị", "don vi", "dv"]) or ""
            thanhtich = get_val(raw_row, ["thanhtich", "thành tích", "thanh tich"]) or ""

            # chuyển namsinh thành int nếu có
            namsinh_val = None
            try:
                if namsinh is not None and str(namsinh).strip() != "":
                    # handle float strings like "1990.0"
                    namsinh_val = int(float(namsinh))
            except Exception:
                namsinh_val = None

            # chuẩn capbac (loại bỏ .0 nếu có)
            if capbac is not None:
                capbac = str(capbac).strip()
                if re.match(r"^\d+\.0+$", capbac):
                    capbac = str(int(float(capbac)))

            # kiểm tra tồn tại -> cập nhật hoặc thêm mới
            s = HoSoVoSinh.query.filter_by(vocotruyenid=code).first()
            trinhdo = get_trinh_do(capbac)
            maudai_rel = get_mau_dai(capbac, mode="rel") if capbac else None

            if s:
                # update các trường nếu trong file có giá trị
                s.hovaten = str(hovaten).strip() or s.hovaten
                s.namsinh = namsinh_val or s.namsinh
                s.gioitinh = gioitinh or s.gioitinh
                s.capbac = capbac or s.capbac
                s.trinhdo = trinhdo or s.trinhdo
                s.maudai = maudai_rel or s.maudai
                s.donvi = donvi or s.donvi
                s.thanhtich = thanhtich or s.thanhtich
            else:
                new_s = HoSoVoSinh(
                    vocotruyenid=code,
                    hovaten=str(hovaten).strip() or "Chưa đặt tên",
                    namsinh=namsinh_val,
                    gioitinh=gioitinh or "",
                    donvi=donvi or "",
                    capbac=capbac or "",
                    trinhdo=trinhdo,
                    maudai=maudai_rel,
                    thanhtich=thanhtich or "",
                )
                db.session.add(new_s)

        db.session.commit()
        flash("Import thành công!")
        return redirect(url_for("admin_students"))

    return render_template("admin/import_students.html")


# ====== Export võ sinh (có ảnh màu đai từ static) ======
@app.route("/admin/students/export")
def admin_students_export():
    if session.get("role") != "admin":
        return redirect(url_for("home"))

    students = HoSoVoSinh.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách võ sinh"

    headers = ["Mã số", "Họ tên", "Năm sinh", "Giới tính", "Cấp bậc", "Trình độ", "Màu đai", "Đơn vị", "Thành tích"]
    ws.append(headers)

    row_num = 2
    for s in students:
        ws.cell(row=row_num, column=1, value=s.vocotruyenid)
        ws.cell(row=row_num, column=2, value=s.hovaten)
        ws.cell(row=row_num, column=3, value=s.namsinh)
        ws.cell(row=row_num, column=4, value=s.gioitinh)
        ws.cell(row=row_num, column=5, value=s.capbac)
        ws.cell(row=row_num, column=6, value=get_trinh_do(s.capbac))
        ws.cell(row=row_num, column=8, value=s.donvi)
        ws.cell(row=row_num, column=9, value=s.thanhtich)

        # Lấy path tuyệt đối cho file ảnh trong thư mục static
        img_path = None
        try:
            img_path = get_mau_dai(s.capbac, mode="path") if s.capbac else None
        except Exception:
            img_path = None

        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                img.width, img.height = 40, 20
                ws.add_image(img, f"G{row_num}")
            except Exception:
                # bỏ qua nếu lỗi chèn ảnh
                pass

        row_num += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name="students_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/students/delete/<vocotruyenid>", methods=["GET"])
def admin_students_delete(vocotruyenid):
    if session.get("role") != "admin":
        return redirect(url_for("home"))

    student = HoSoVoSinh.query.get(vocotruyenid)
    if student:
        db.session.delete(student)
        db.session.commit()
    return redirect(url_for("admin_students"))


@app.route("/admin/students/update/<vocotruyenid>", methods=["POST"])
def admin_students_update(vocotruyenid):
    if session.get("role") != "admin":
        return redirect(url_for("home"))

    student = HoSoVoSinh.query.get(vocotruyenid)
    if not student:
        return redirect(url_for("admin_students"))

    student.hovaten = request.form.get("hovaten") or student.hovaten
    namsinh_val = request.form.get("namsinh")
    if namsinh_val:
        try:
            student.namsinh = int(namsinh_val)
        except Exception:
            pass
    student.gioitinh = request.form.get("gioitinh") or student.gioitinh

    capbac_val = request.form.get("capbac")
    student.capbac = capbac_val or student.capbac

    if student.capbac:
        student.trinhdo = get_trinh_do(student.capbac)
        student.maudai = get_mau_dai(student.capbac, mode="rel")
    else:
        student.trinhdo = "Chưa xếp"
        student.maudai = None

    student.donvi = request.form.get("donvi") or student.donvi
    student.thanhtich = request.form.get("thanhtich") or student.thanhtich

    db.session.commit()
    return redirect(url_for("admin_students"))


# ---------------- MAIN ----------------
if __name__ == "__main__":
    # debug=True chỉ dùng khi dev
    app.run(debug=True, host="0.0.0.0", port=5000)
